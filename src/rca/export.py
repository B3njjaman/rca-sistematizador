"""Etapa 5: consolidación y export (xlsx, csv, json, html)."""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd

from .schema import COLUMNAS_EXPORT, Exigencia


def _dataframe(exigencias: list[Exigencia]) -> pd.DataFrame:
    filas = [ex.to_row() for ex in exigencias]
    df = pd.DataFrame(filas)
    # asegura todas las columnas y el orden, aunque no haya filas
    for col in COLUMNAS_EXPORT:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNAS_EXPORT]


def exportar(exigencias: list[Exigencia], out_dir: Path, nombre: str = "matriz_exigencias") -> dict:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    df = _dataframe(exigencias)

    ruta_json = out_dir / f"{nombre}.json"
    ruta_csv = out_dir / f"{nombre}.csv"
    ruta_xlsx = out_dir / f"{nombre}.xlsx"
    ruta_html = out_dir / f"{nombre}.html"

    ruta_json.write_text(
        json.dumps([e.model_dump() for e in exigencias], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    df.to_csv(ruta_csv, index=False, encoding="utf-8-sig")  # utf-8-sig -> Excel lee tildes ok
    _to_excel(df, ruta_xlsx)
    _to_html(df, ruta_html)

    return {"json": ruta_json, "csv": ruta_csv, "xlsx": ruta_xlsx, "html": ruta_html, "filas": len(df)}


def _to_excel(df: pd.DataFrame, ruta: Path) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Matriz RCA")
        ws = writer.sheets["Matriz RCA"]
        encabezado = Font(bold=True, color="FFFFFF")
        relleno = PatternFill("solid", fgColor="1F4E78")
        for j, _ in enumerate(df.columns, start=1):
            celda = ws.cell(row=1, column=j)
            celda.font = encabezado
            celda.fill = relleno
            celda.alignment = Alignment(vertical="center", wrap_text=True)
        # anchos: columnas de texto largo más anchas
        anchas = {"Transcripción Literal", "Verificadores Propuestos", "Antecedentes Complementarios", "Nombre Exigencia"}
        for j, col in enumerate(df.columns, start=1):
            letra = get_column_letter(j)
            ws.column_dimensions[letra].width = 60 if col in anchas else 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def _to_html(df: pd.DataFrame, ruta: Path) -> None:
    estilo = """
    <style>
      body{font-family:system-ui,Arial,sans-serif;margin:1.5rem;}
      h1{font-size:1.2rem;}
      table{border-collapse:collapse;font-size:.8rem;}
      th{background:#1F4E78;color:#fff;position:sticky;top:0;padding:6px;text-align:left;}
      td{border:1px solid #ddd;padding:6px;vertical-align:top;max-width:420px;}
      tr:nth-child(even){background:#f7f9fc;}
    </style>
    """
    tabla = df.to_html(index=False, escape=True, na_rep="")
    html = f"<!doctype html><html lang='es'><head><meta charset='utf-8'>{estilo}</head><body>" \
           f"<h1>Matriz de exigencias ({len(df)} filas)</h1>{tabla}</body></html>"
    ruta.write_text(html, encoding="utf-8")
